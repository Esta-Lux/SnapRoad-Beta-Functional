"""
SnapRoad Admin API Routes
All endpoints use the Supabase DAO layer (supabase_service.py).
"""
from pydantic import BaseModel, Field
from fastapi import APIRouter, Body, Depends, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Annotated, Optional, List
from datetime import datetime, timedelta, timezone
import json
import logging
import io
import secrets
from middleware.auth import require_admin

from models.schemas import (
    AdminOfferCreate, PricingUpdate, OfferImport, BulkOfferUpload,
    BoostCalculate, BoostCreate, AnalyticsEvent,
)
from services.offer_utils import calculate_auto_gems, calculate_free_discount
from services.offer_categories import attach_offer_category_fields
from services.cache import cache_delete
from services.fee_calculator import list_monthly_fee_summaries, get_partner_fee_history
from services.offer_analytics import (
    summarize_offer_analytics,
    recent_offer_feed,
    map_data_for_day,
    today_realtime_summary,
    record_offer_event,
)
from services.telemetry_service import telemetry_service
from services.supabase_service import (
    sb_list_profiles, sb_get_profile, sb_get_profile_raw, sb_update_profile,
    sb_update_partner,
    compute_extended_promotion_until_iso,
    sb_get_profile_promotion_until_raw,
    sb_get_partner_promotion_until_raw,
    sb_suspend_profile, sb_activate_profile, sb_delete_profile,
    sb_get_partners, sb_get_partner, sb_get_partners_plan_fields_by_ids, sb_create_partner,
    sb_delete_partner,
    sb_get_partner_locations_for_admin_map,
    sb_get_offers, sb_create_offer, sb_update_offer, sb_delete_offer,
    sb_get_boosts, sb_create_boost, sb_cancel_boost,
    sb_get_redemption_stats,
    sb_get_challenges,
    sb_get_incidents, sb_update_incident,
    sb_get_admin_notifications, sb_create_notification,
    sb_mark_notification_read,
    sb_get_campaigns, sb_create_campaign, sb_update_campaign, sb_delete_campaign,
    sb_get_rewards, sb_create_reward, sb_update_reward, sb_delete_reward,
    sb_get_audit_logs, sb_create_audit_log,
    sb_get_settings, sb_update_setting,
    sb_get_legal_documents, sb_create_legal_document, sb_update_legal_document,
    sb_get_referral_analytics,
    sb_get_finance_summary,
    sb_get_platform_stats, sb_get_trips_stats,
    test_connection,
    sb_get_concerns, sb_update_concern_status, sb_get_concerns_count_by_status,
    sb_get_app_config, sb_update_app_config,
    sb_get_app_config_with_meta,
    sb_get_road_reports_for_admin_map,
    sb_get_road_reports_admin_list,
    sb_update_road_report_moderation,
    sb_get_live_users,
)

from services.outbound_email import send_html_email, promotion_email_html

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api", tags=["Admin"], dependencies=[Depends(require_admin)])
admin_platform_router = APIRouter()
admin_concerns_router = APIRouter()
admin_config_router = APIRouter()
admin_incidents_router = APIRouter()
admin_offers_router = APIRouter()
admin_partners_router = APIRouter()
admin_campaigns_router = APIRouter()
admin_rewards_router = APIRouter()
admin_users_router = APIRouter()
admin_pricing_router = APIRouter()
admin_boosts_router = APIRouter()
admin_realtime_router = APIRouter()
admin_photo_reports_router = APIRouter()
admin_ops_router = APIRouter()

AdminUser = Annotated[dict, Depends(require_admin)]

class AdminPhotoRejectBody(BaseModel):
    review_notes: Optional[str] = None


class ConcernStatusBody(BaseModel):
    """POST JSON: { "status": "open" | "in_progress" | "resolved" | "closed" }."""

    status: str = Field(..., min_length=1, max_length=32)


BOOST_PRICING_ADMIN = {
    "base_daily_cost": 25, "additional_day_cost": 20,
    "base_reach": 100, "base_reach_cost": 5,
    "reach_increment": 100, "reach_increment_cost": 10,
}


# ==================== PLATFORM STATS ====================

@admin_platform_router.get("/admin/stats")
def get_admin_stats():
    stats = sb_get_platform_stats()
    trip_stats = sb_get_trips_stats()
    redemption_stats = sb_get_redemption_stats()
    finance = sb_get_finance_summary()
    open_concerns = sb_get_concerns_count_by_status("open")
    pending_incidents = len(sb_get_incidents(status="pending", limit=500))
    if not stats:
        stats = {}
    data = {
        **stats,
        "active_today": stats.get("active_today", len(sb_get_live_users())),
        "trips_today": stats.get("trips_today", trip_stats.get("total_trips", 0)),
        "total_miles": stats.get("total_miles", 0),
        "open_concerns": open_concerns,
        "pending_incidents": pending_incidents,
        "offers_redeemed": stats.get("total_redemptions", redemption_stats.get("total", 0)),
        "revenue": finance.get("total_mrr", 0),
        "user_growth": stats.get("user_growth", "+0%"),
    }
    return {"success": True, "source": "supabase", "data": data}


@admin_platform_router.get("/admin/wallet-ledger")
def admin_wallet_ledger(
    user_id: Annotated[Optional[str], Query(description="Filter by driver user_id")] = None,
    tx_type: Annotated[Optional[str], Query(description="Filter by tx_type e.g. trip_drive, offer_redeem")] = None,
    limit: Annotated[int, Query(ge=1, le=500)] = 200,
):
    """Recent rows from `wallet_transactions` for ops review; optional filter by driver user_id or tx_type."""
    from database import get_supabase

    sb = get_supabase()
    try:
        q = (
            sb.table("wallet_transactions")
            .select(
                "id,user_id,tx_type,direction,amount,balance_before,balance_after,"
                "reference_type,reference_id,metadata,status,created_at"
            )
            .order("created_at", desc=True)
            .limit(limit)
        )
        if user_id:
            q = q.eq("user_id", user_id)
        if tx_type and str(tx_type).strip():
            q = q.eq("tx_type", str(tx_type).strip())
        res = q.execute()
        return {"success": True, "data": res.data or []}
    except Exception as e:
        logger.warning("admin wallet-ledger: %s", e)
        return {"success": False, "message": "Could not load wallet ledger", "data": []}


@admin_platform_router.get("/admin/redemptions")
def admin_redemptions_list(
    user_id: Annotated[Optional[str], Query(description="Filter by driver user_id")] = None,
    status: Annotated[Optional[str], Query(description="Redemption status e.g. verified")] = None,
    limit: Annotated[int, Query(ge=1, le=500)] = 150,
):
    """Monitor offer redemptions across partners (service role)."""
    from database import get_supabase

    sb = get_supabase()
    try:
        q = sb.table("redemptions").select("*").order("created_at", desc=True).limit(limit)
        if user_id:
            q = q.eq("user_id", user_id)
        if status and str(status).strip():
            q = q.eq("status", str(status).strip())
        res = q.execute()
        return {"success": True, "data": res.data or []}
    except Exception as e:
        logger.warning("admin redemptions: %s", e)
        return {"success": False, "message": "Could not load redemptions", "data": []}


class AdminOfferStatusBody(BaseModel):
    """Set offer visibility / lifecycle for moderation."""

    status: str = Field(..., min_length=3, max_length=32)


@admin_offers_router.post("/admin/offers/{offer_id}/status")
def admin_set_offer_status(offer_id: str, body: AdminOfferStatusBody):
    """
    Approve or pause offers: active | inactive | paused.
    `rejected` is stored as inactive for drivers (not shown as live offers).
    """
    raw = str(body.status or "").strip().lower()
    if raw == "reject":
        raw = "rejected"
    if raw not in ("active", "inactive", "paused", "rejected"):
        return {"success": False, "message": "status must be active, inactive, paused, or rejected"}
    stored = "inactive" if raw == "rejected" else raw
    ok = sb_update_offer(offer_id, {"status": stored})
    if ok:
        sb_create_audit_log("OFFER_STATUS", "admin", offer_id, f"Offer status set to {stored}")
        return {"success": True, "message": "Offer updated", "data": {"status": stored}}
    return {"success": False, "message": "Could not update offer"}


# ==================== CONCERNS (admin) ====================

@admin_concerns_router.get("/admin/concerns")
def get_admin_concerns(
    limit: Annotated[int, Query(ge=1, le=200)] = 50,
    severity: Annotated[Optional[str], Query()] = None,
    status: Annotated[Optional[str], Query()] = None,
):
    concerns = sb_get_concerns(limit=limit, severity=severity, status=status)
    return {"success": True, "data": {"concerns": concerns, "total": len(concerns)}}


@admin_concerns_router.post("/admin/concerns/{concern_id}/status")
def update_concern_status(concern_id: str, body: ConcernStatusBody):
    status = body.status.strip().lower()
    if status not in ("open", "in_progress", "resolved", "closed"):
        return {"success": False, "message": "Invalid status"}
    ok = sb_update_concern_status(concern_id, status)
    if ok:
        return {"success": True, "message": "Updated"}
    return {"success": False, "message": "Update failed"}


# ==================== LIVE USERS ====================

@admin_platform_router.get("/admin/live-users")
def get_live_users():
    users = sb_get_live_users()
    return {"success": True, "data": {"users": users}}


# ==================== TELEMETRY (aggregate app / API usage) ====================

_SKIP_USAGE_PREFIXES = (
    "/docs",
    "/redoc",
    "/openapi.json",
    "/api/admin/monitor",
    "/api/ws/",
    "/favicon.ico",
)


@admin_platform_router.get("/admin/telemetry/app-usage")
def get_admin_app_usage_telemetry(limit: Annotated[int, Query(ge=50, le=500)] = 500):
    """
    Summarize recent HTTP telemetry into API area counts (driver/partner flows proxy).

    Privacy: no per-user identity in telemetry events today—this is aggregate traffic only.
    """
    from collections import Counter

    events = telemetry_service.snapshot(limit=limit)
    prefix_counts: Counter[str] = Counter()
    path_counts: Counter[str] = Counter()
    analyzed = 0

    for e in events:
        path = (e.get("path") or "").split("?")[0]
        if not path.startswith("/api/"):
            continue
        if any(path.startswith(p) for p in _SKIP_USAGE_PREFIXES):
            continue
        analyzed += 1
        path_counts[path] += 1
        parts = [p for p in path.split("/") if p]
        if len(parts) >= 2:
            prefix = f"/{parts[0]}/{parts[1]}"
        else:
            prefix = path or "/"
        prefix_counts[prefix] += 1

    top_prefixes = [{"prefix": k, "count": v} for k, v in prefix_counts.most_common(24)]
    top_paths = [{"path": k, "count": v} for k, v in path_counts.most_common(30)]

    return {
        "success": True,
        "data": {
            "events_in_buffer": len(events),
            "api_events_counted": analyzed,
            "top_prefixes": top_prefixes,
            "top_paths": top_paths,
        },
    }


# ==================== HEALTH ====================

def _admin_health_supabase(results: dict) -> None:
    import time

    from database import get_supabase

    try:
        start = time.time()
        supabase = get_supabase()
        supabase.table("app_config").select("key").limit(1).execute()
        results["database"] = "healthy"
        results["db_latency"] = round((time.time() - start) * 1000)
    except Exception as e:
        results["database"] = "down"
        results["db_error"] = str(e)


async def _admin_health_google_maps(results: dict) -> None:
    import os
    import time

    import httpx

    key = os.environ.get("GOOGLE_MAPS_API_KEY") or os.environ.get("GOOGLE_PLACES_API_KEY")
    try:
        start = time.time()
        async with httpx.AsyncClient(timeout=5) as client:
            r = await client.get(
                "https://maps.googleapis.com/maps/api/geocode/json",
                params={"address": "Columbus OH", "key": key or ""},
            )
        results["google_maps"] = "healthy" if r.status_code == 200 else "degraded"
        results["maps_latency"] = round((time.time() - start) * 1000)
    except Exception:
        results["google_maps"] = "down"


async def _admin_health_ohgo(results: dict) -> None:
    import os
    import time

    import httpx

    ohgo_key = os.environ.get("VITE_OHGO_API_KEY") or os.environ.get("OHGO_API_KEY")
    try:
        start = time.time()
        async with httpx.AsyncClient(timeout=5) as client:
            r = await client.get(
                "https://publicapi.ohgo.com/api/v1/cameras",
                params={"api-key": ohgo_key or "", "page-size": "1"},
            )
        results["ohgo"] = "healthy" if r.status_code == 200 else "degraded"
        results["ohgo_latency"] = round((time.time() - start) * 1000)
    except Exception:
        results["ohgo"] = "down"


async def _admin_health_llm(results: dict) -> None:
    import os
    import time

    import httpx

    async def ping_models(url: str, key: str) -> tuple[str, int]:
        if not key.strip():
            return "not_configured", 0
        try:
            start = time.time()
            async with httpx.AsyncClient(timeout=8) as client:
                r = await client.get(url, headers={"Authorization": f"Bearer {key}"})
            lat = round((time.time() - start) * 1000)
            status = "healthy" if r.status_code == 200 else "degraded"
            return status, lat
        except Exception:
            return "down", 0

    okey = (os.environ.get("OPENAI_API_KEY") or os.environ.get("OPENAI_KEY") or "").strip()
    nkey = (os.environ.get("NVIDIA_API_KEY") or "").strip()
    o_url = "https://api.openai.com/v1/models"
    n_base = (os.environ.get("NVIDIA_API_BASE") or "https://integrate.api.nvidia.com/v1").strip().rstrip("/")
    n_url = f"{n_base}/models"

    o_st, o_lat = await ping_models(o_url, okey)
    n_st, n_lat = await ping_models(n_url, nkey)

    results["llm_openai"] = o_st if okey else "not_configured"
    results["llm_openai_latency"] = o_lat
    results["llm_nvidia"] = n_st if nkey else "not_configured"
    results["llm_nvidia_latency"] = n_lat

    if okey and nkey:
        results["llm_provider"] = "openai_primary_nvidia_fallback"
    elif okey:
        results["llm_provider"] = "openai"
    elif nkey:
        results["llm_provider"] = "nvidia"
    else:
        results["llm_provider"] = "none"

    if results["llm_openai"] == "healthy":
        results["llm"] = "healthy"
        results["llm_latency"] = results["llm_openai_latency"]
    elif results["llm_nvidia"] == "healthy":
        results["llm"] = "healthy"
        results["llm_latency"] = results["llm_nvidia_latency"]
    elif results["llm_openai"] in ("degraded", "down") and results["llm_nvidia"] == "healthy":
        results["llm"] = "degraded"
        results["llm_latency"] = results["llm_nvidia_latency"]
    else:
        results["llm"] = "down"
        results["llm_latency"] = max(o_lat, n_lat)

    results["openai"] = results["llm_openai"]
    results["openai_latency"] = results["llm_openai_latency"]


async def _admin_health_realtime(results: dict) -> None:
    import os

    import httpx

    supabase_url = os.environ.get("SUPABASE_URL")
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            r = await client.get(f"{supabase_url or ''}/rest/v1/")
        results["realtime"] = "healthy" if r.status_code < 500 else "degraded"
    except Exception:
        results["realtime"] = "down"


@admin_platform_router.get("/admin/health")
async def get_admin_health():
    results: dict = {"api": "healthy"}
    _admin_health_supabase(results)
    await _admin_health_google_maps(results)
    await _admin_health_ohgo(results)
    await _admin_health_llm(results)
    await _admin_health_realtime(results)
    return {"success": True, "data": results}


# ==================== APP CONFIG (admin) ====================

@admin_config_router.get("/admin/config")
def get_admin_config():
    config = sb_get_app_config()
    return {"success": True, "data": config}


@admin_config_router.post("/admin/config")
def update_admin_config(body: dict, user: AdminUser):
    """Apply key/value pairs to app_config. Reserved keys: _reason, reason (ops runbook text for audit)."""
    updated_by = user.get("user_id") if user else None
    payload = dict(body) if isinstance(body, dict) else {}
    reason = payload.pop("_reason", None) or payload.pop("reason", None)
    payload.pop("_previous_snapshot", None)  # optional client echo; not persisted

    patch = {k: v for k, v in payload.items() if isinstance(k, str) and not k.startswith("_")}
    if not patch:
        return {"success": False, "message": "No config keys to update", "data": sb_get_app_config()}

    before = sb_get_app_config()
    changes = {}
    for key, value in patch.items():
        changes[key] = {"from": before.get(key), "to": value}
        sb_update_app_config(key, value, updated_by=updated_by)

    cache_delete("app_config_public")
    try:
        from services.runtime_config import invalidate_runtime_config_cache

        invalidate_runtime_config_cache()
    except Exception as e:
        logger.warning("failed to invalidate runtime config cache: %s", e)

    keys_sorted = sorted(patch.keys())
    audit_obj = {
        "keys": keys_sorted,
        "changes": changes,
        "reason": (str(reason).strip() or None),
        "updated_by": str(updated_by or ""),
    }
    detail = json.dumps(audit_obj, default=str)[:12000]
    sb_create_audit_log(
        "APP_CONFIG_UPDATED",
        "admin",
        str(updated_by or ""),
        detail,
    )
    return {"success": True, "data": sb_get_app_config()}


@admin_config_router.get("/admin/config/detailed")
def get_admin_config_detailed():
    """Config values plus per-key updated_at / updated_by for ops audit trail in UI."""
    cfg, meta = sb_get_app_config_with_meta()
    return {"success": True, "data": {"config": cfg, "meta": meta}}


@admin_config_router.get("/admin/map/road-reports")
def get_admin_map_road_reports(limit: Annotated[int, Query(ge=1, le=800)] = 400):
    reports = sb_get_road_reports_for_admin_map(limit=limit)
    return {"success": True, "data": {"reports": reports}}


@admin_config_router.get("/admin/map/partner-locations")
def get_admin_map_partner_locations(limit: Annotated[int, Query(ge=1, le=1000)] = 500):
    locations = sb_get_partner_locations_for_admin_map(limit=limit)
    return {"success": True, "data": {"locations": locations}}


# ==================== ANALYTICS ====================

@admin_platform_router.get("/admin/analytics")
def get_admin_analytics():
    stats = sb_get_platform_stats()
    trip_stats = sb_get_trips_stats()
    redemption_stats = sb_get_redemption_stats()
    finance = sb_get_finance_summary()

    incidents_pending = len(sb_get_incidents(status="pending", limit=200))

    return {
        "success": True,
        "data": {
            "summary": {
                "total_users": stats.get("total_users", 0),
                "premium_users": stats.get("premium_users", 0),
                "avg_safety_score": stats.get("avg_safety_score", 0),
                "total_partners": stats.get("total_partners", 0),
                "active_partners": stats.get("active_partners", 0),
                "total_offers": stats.get("total_offers", 0),
                "total_trips": trip_stats.get("total_trips", 0),
                "total_redemptions": redemption_stats.get("total", 0),
                "total_gems": stats.get("total_gems", 0),
                "total_mrr": finance.get("total_mrr", 0),
            },
            "queues": {
                "incident_review": incidents_pending,
                "consent_pending": 0,
                "fraud_flags": 0,
            },
        },
    }


# ==================== REFERRAL ANALYTICS ====================

@admin_platform_router.get("/admin/referral-analytics")
def get_referral_analytics():
    data = sb_get_referral_analytics()
    return {"success": True, "data": data}


# ==================== FINANCE ====================

@admin_platform_router.get("/admin/finance")
def get_finance_data():
    summary = sb_get_finance_summary()
    return {"success": True, "data": {"summary": summary}}


# ==================== NOTIFICATIONS ====================

@admin_ops_router.get("/admin/notifications")
def get_notifications(limit: Annotated[int, Query(ge=1, le=100)] = 50):
    data = sb_get_admin_notifications(limit=limit)
    return {"success": True, "data": data}


@admin_ops_router.post("/admin/notifications")
def create_notification_endpoint(notification_data: dict):
    result = sb_create_notification(notification_data)
    if result:
        return {"success": True, "message": "Notification created", "data": result}
    return {"success": False, "message": "Failed to create notification"}


@admin_ops_router.patch("/admin/notifications/{notification_id}/read")
def mark_notification_read(notification_id: str):
    success = sb_mark_notification_read(notification_id)
    if success:
        return {"success": True, "message": "Notification marked as read"}
    return {"success": False, "message": "Failed to mark notification as read"}


# ==================== LEGAL DOCUMENTS ====================

@admin_ops_router.get("/admin/legal-documents")
def get_legal_documents():
    data = sb_get_legal_documents()
    return {"success": True, "data": data}


_LEGAL_DOC_KEYS = frozenset(
    {"name", "slug", "type", "status", "version", "description", "content", "is_required"}
)


@admin_ops_router.post("/admin/legal-documents")
def create_legal_document(doc_data: dict):
    body = {k: v for k, v in doc_data.items() if k in _LEGAL_DOC_KEYS}
    created = sb_create_legal_document(body)
    if created:
        sb_create_audit_log(
            "LEGAL_DOC_CREATED",
            "admin",
            str(created.get("id", "")),
            "Created legal document",
        )
        return {"success": True, "data": created, "message": "Document created"}
    return {
        "success": False,
        "message": "Failed to create document (name and type are required)",
    }


@admin_ops_router.put("/admin/legal-documents/{doc_id}")
def update_legal_document(doc_id: str, doc_data: dict):
    body = {k: v for k, v in doc_data.items() if k in _LEGAL_DOC_KEYS}
    body["last_updated"] = datetime.now(timezone.utc).isoformat()
    success = sb_update_legal_document(doc_id, body)
    if success:
        sb_create_audit_log("LEGAL_DOC_UPDATED", "admin", doc_id, "Updated legal document")
        return {"success": True, "message": "Document updated"}
    return {"success": False, "message": "Failed to update document"}


@admin_ops_router.post("/admin/legal-documents/seed-defaults")
def seed_default_legal_documents_endpoint(payload: Optional[dict] = Body(default=None)):
    """
    One-shot seeder for the default Terms of Service + Privacy Policy.

    Idempotent: existing rows (matched on `slug`) are left untouched unless
    the request body sets `{"force": true}`, which overwrites their `content`
    from the on-disk seed files. Safe to call after every deploy.
    """
    from services.legal_seed import seed_default_legal_documents
    force = bool((payload or {}).get("force"))
    result = seed_default_legal_documents(force=force)
    sb_create_audit_log(
        "LEGAL_DOC_SEEDED",
        "admin",
        "legal_documents",
        f"Seed defaults (force={force}): inserted={result['inserted']}, updated={result['updated']}, skipped={result['skipped']}",
    )
    has_errors = bool(result.get("errors"))
    return {
        "success": not has_errors,
        "message": (
            "Default legal documents already present"
            if not result["inserted"] and not result["updated"] and not has_errors
            else "Default legal documents seeded"
        ),
        "data": result,
    }


# ==================== SETTINGS ====================

@admin_ops_router.get("/admin/settings")
def get_settings():
    data = sb_get_settings()
    return {"success": True, "data": data}


@admin_ops_router.post("/admin/settings")
def update_settings(settings_data: dict):
    for key, value in settings_data.items():
        sb_update_setting(key, value)
    sb_create_audit_log("SETTINGS_UPDATED", "admin", "platform_settings", "Updated platform settings")
    return {"success": True, "message": "Settings updated successfully", "data": settings_data}


# ==================== AUDIT LOG ====================

@admin_ops_router.get("/admin/audit-log")
def get_audit_log(limit: Annotated[int, Query(ge=1, le=100)] = 50):
    data = sb_get_audit_logs(limit=limit)
    return {"success": True, "data": data}


# ==================== INCIDENTS ====================

def _road_report_severity(report_type: str) -> str:
    if report_type in ("accident", "crash", "closure"):
        return "high"
    if report_type in ("pothole",):
        return "low"
    return "medium"


def _road_report_row_to_admin_item(row: dict) -> dict:
    """Align driver map reports (road_reports) with admin Incidents tab shape."""
    t = str(row.get("type") or "report")
    sev = _road_report_severity(t)
    lat, lng = row.get("lat"), row.get("lng")
    loc = ""
    try:
        if lat is not None and lng is not None:
            loc = f"{float(lat):.5f}, {float(lng):.5f}"
    except (TypeError, ValueError):
        loc = ""
    mod = str(row.get("moderation_status") or "approved").lower()
    # AIModerationTab maps `status` via mapStatus (pending → new). Row `status` is row lifecycle, not moderation.
    queue_status = mod if mod in ("pending", "approved", "rejected") else "approved"
    return {
        "id": str(row.get("id")),
        "type": t,
        "description": row.get("description") or "",
        "location": loc,
        "severity": sev,
        "status": queue_status,
        "moderation_status": mod,
        "created_at": row.get("created_at"),
        "reported_by": str(row.get("user_id") or ""),
        "source": "road_reports",
        "image_url": None,
        # AIModerationTab filters by confidence (default threshold 80); driver reports have no ML score.
        "confidence": 0.9,
    }


@admin_incidents_router.get("/admin/incidents")
def get_incidents(
    limit: Annotated[int, Query(ge=1, le=200)] = 100,
    status: Annotated[Optional[str], Query()] = None,
):
    legacy = sb_get_incidents(status=status, limit=limit)
    road_rows = sb_get_road_reports_admin_list(min(limit, 120))
    road_items = [_road_report_row_to_admin_item(r) for r in road_rows]
    merged = road_items + list(legacy)

    def sort_key(item: dict):
        return str(item.get("created_at") or "")

    merged.sort(key=sort_key, reverse=True)
    return {"success": True, "data": merged[:limit]}


@admin_incidents_router.post("/admin/incidents/{incident_id}/moderate")
async def moderate_incident(incident_id: str, outcome: Annotated[str, Body(..., embed=True)]):
    if outcome not in ["approved", "rejected"]:
        return {"success": False, "message": "Invalid outcome. Must be 'approved' or 'rejected'"}

    if sb_update_road_report_moderation(incident_id, outcome):
        sb_create_audit_log("INCIDENT_MODERATED", "admin", incident_id, f"Road report {outcome}")
        try:
            from services.websocket_manager import ws_manager
            await ws_manager.broadcast_moderation_update(incident_id, outcome)
        except Exception as e:
            logger.warning(f"Could not broadcast moderation update: {e}")
        return {"success": True, "message": "Update saved"}

    success = sb_update_incident(incident_id, {
        "status": outcome,
        "moderated_by": "admin",
        "moderated_at": datetime.now().isoformat(),
    })

    if success:
        sb_create_audit_log("INCIDENT_MODERATED", "admin", incident_id, f"Incident {outcome}")
        try:
            from services.websocket_manager import ws_manager
            await ws_manager.broadcast_moderation_update(incident_id, outcome)
        except Exception as e:
            logger.warning(f"Could not broadcast moderation update: {e}")
        return {"success": True, "message": "Update saved"}
    return {"success": False, "message": "Could not update that report. Try refreshing the list."}


@admin_incidents_router.get("/admin/incidents/moderated")
def get_moderated_incidents(limit: Annotated[int, Query(ge=1, le=100)] = 100):
    approved = sb_get_incidents(status="approved", limit=limit)
    rejected = sb_get_incidents(status="rejected", limit=limit)
    return {"success": True, "data": approved + rejected, "total": len(approved) + len(rejected)}


# ==================== OFFERS CRUD ====================

@admin_offers_router.get("/admin/offers")
def get_offers(
    status: Annotated[str, Query()] = "all",
    limit: Annotated[int, Query(ge=1, le=100)] = 100,
):
    data = sb_get_offers(status=status, limit=limit)
    for row in data:
        attach_offer_category_fields(row)
    return {"success": True, "data": data}


@admin_offers_router.post("/admin/offers")
def create_offer(offer_data: dict):
    offer_data.setdefault("status", "active")
    offer_data.setdefault("is_admin_offer", True)
    offer_data.setdefault("created_by", "admin")
    result = sb_create_offer(offer_data)
    if result:
        sb_create_audit_log("OFFER_CREATED", "admin", result.get("id", ""), f"Created offer: {offer_data.get('title', '')}")
        return {"success": True, "message": "Offer created successfully", "data": result}
    return {"success": False, "message": "Failed to create offer"}

def _optional_fields(obj, keys_map: dict) -> dict:
    """Return a dict of non-None optional fields from *obj*."""
    return {k: getattr(obj, v) for k, v in keys_map.items() if getattr(obj, v) is not None}


_OFFER_OPTIONAL_KEYS = {
    "offer_url": "offer_url",
    "original_price": "original_price",
    "affiliate_tracking_url": "affiliate_tracking_url",
    "external_id": "external_id",
}


@admin_offers_router.post("/admin/offers/create")
def admin_create_offer(offer: AdminOfferCreate):
    auto_gems = offer.base_gems if offer.base_gems is not None else calculate_auto_gems(offer.discount_percent, offer.is_free_item)
    premium_discount = offer.discount_percent
    free_discount = calculate_free_discount(premium_discount)

    title_val = (offer.title or "").strip() or (
        (offer.description[:60] if offer.description else "") or "Admin Offer"
    )
    img_val = (offer.image_url or offer.image_id or None)
    img_val = str(img_val).strip() if img_val else None

    data = {
        "business_name": offer.business_name,
        "business_type": offer.business_type,
        "description": offer.description,
        "discount_percent": premium_discount,
        "premium_discount_percent": premium_discount,
        "free_discount_percent": free_discount,
        "is_free_item": offer.is_free_item,
        "base_gems": auto_gems,
        "lat": offer.lat,
        "lng": offer.lng,
        "is_admin_offer": True,
        "created_by": "admin",
        "status": "active",
        "title": title_val,
        "offer_source": offer.offer_source,
        **_optional_fields(offer, _OFFER_OPTIONAL_KEYS),
    }
    if img_val:
        data["image_url"] = img_val
    if offer.expires_hours:
        data["expires_at"] = (datetime.now() + timedelta(hours=offer.expires_hours)).isoformat()

    result = sb_create_offer(data)
    if result:
        sb_create_audit_log("OFFER_CREATED", "admin", result.get("id", ""), f"Created offer for {offer.business_name}")
        return {"success": True, "message": f"Offer created for {offer.business_name}", "data": result}
    return {"success": False, "message": "Failed to create offer"}


_BULK_OPTIONAL_KEYS = {
    "original_price": "original_price",
    "affiliate_tracking_url": "affiliate_tracking_url",
    "external_id": "external_id",
}


def _build_bulk_offer(item) -> dict:
    """Build a single offer dict from a BulkOfferUpload item."""
    auto_gems = item.base_gems if item.base_gems is not None else calculate_auto_gems(item.discount_percent, item.is_free_item)
    premium_discount = item.discount_percent
    return {
        "business_name": item.business_name,
        "business_type": item.business_type,
        "description": item.description,
        "discount_percent": premium_discount,
        "premium_discount_percent": premium_discount,
        "free_discount_percent": calculate_free_discount(premium_discount),
        "is_free_item": item.is_free_item,
        "base_gems": auto_gems,
        "address": item.address,
        "lat": item.lat or 39.9612,
        "lng": item.lng or -82.9988,
        "offer_url": item.offer_url,
        "is_admin_offer": True,
        "created_by": "admin",
        "status": "active",
        "title": item.description[:60] if item.description else "Bulk Offer",
        "expires_at": (datetime.now() + timedelta(days=item.expires_days)).isoformat() if item.expires_days else None,
        "offer_source": item.offer_source,
        **_optional_fields(item, _BULK_OPTIONAL_KEYS),
    }


@admin_offers_router.post("/admin/offers/bulk")
def admin_bulk_offers(data: BulkOfferUpload):
    created = []
    for item in data.offers:
        result = sb_create_offer(_build_bulk_offer(item))
        if result:
            created.append(result)
    return {"success": True, "message": f"Created {len(created)} offers", "data": {"created_count": len(created), "offers": created}}


def _excel_cell(row: tuple, col_map: dict, col: str, default=""):
    """Get a cell value from an Excel row by column name."""
    idx = col_map.get(col)
    if idx is None or idx >= len(row) or row[idx] is None:
        return default
    return row[idx]


_VALID_SOURCES = {"direct", "groupon", "affiliate", "yelp", "manual"}


def _parse_excel_offer_row(row: tuple, col_map: dict) -> dict:
    """Build an offer dict from a single Excel row. Returns the offer dict or raises on error."""
    get = lambda col, default="": _excel_cell(row, col_map, col, default)

    biz_name = str(get("business_name", "")).strip()
    if not biz_name:
        raise ValueError("missing business_name")

    disc = int(float(get("discount_percent", 0)))
    is_free = str(get("is_free_item", "")).lower() in ("true", "yes", "1")

    source = str(get("source", get("offer_source", "direct"))).strip().lower() or "direct"
    if source not in _VALID_SOURCES:
        source = "direct"

    offer_data = {
        "business_name": biz_name,
        "business_type": str(get("business_type", "retail")),
        "description": str(get("description", "")),
        "title": str(get("title", ""))[:60] or str(get("description", ""))[:60] or "Uploaded Offer",
        "discount_percent": disc,
        "premium_discount_percent": disc,
        "free_discount_percent": calculate_free_discount(disc),
        "is_free_item": is_free,
        "base_gems": calculate_auto_gems(disc, is_free),
        "address": str(get("address", "")),
        "offer_url": str(get("offer_url", "")) or None,
        "lat": float(get("lat", 39.9612)),
        "lng": float(get("lng", -82.9988)),
        "is_admin_offer": True,
        "created_by": "admin_excel",
        "status": "active",
        "expires_at": (datetime.now() + timedelta(days=int(float(get("expires_days", 30))))).isoformat(),
        "offer_source": source,
    }

    orig_price = get("original_price", "")
    if orig_price:
        try:
            offer_data["original_price"] = float(orig_price)
        except (ValueError, TypeError):
            pass
    aff_url = str(get("affiliate_tracking_url", "")).strip()
    if aff_url:
        offer_data["affiliate_tracking_url"] = aff_url
    ext_id = str(get("external_id", "")).strip()
    if ext_id:
        offer_data["external_id"] = ext_id

    return offer_data


def _process_excel_rows(rows: list, col_map: dict) -> tuple[list, list]:
    """Process all data rows from an Excel upload, returning (created, errors)."""
    created = []
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            offer_data = _parse_excel_offer_row(row, col_map)
            result = sb_create_offer(offer_data)
            if result:
                created.append(result)
            else:
                errors.append(f"Row {row_idx}: DB insert failed")
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    return created, errors


@admin_offers_router.post("/admin/offers/upload-excel")
async def upload_excel_offers(file: Annotated[UploadFile, File()]):
    """Parse an Excel (.xlsx) file and create offers. Auto-calculates gems and discounts."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return {"success": False, "message": "Please upload an .xlsx file"}

    try:
        import openpyxl
    except ImportError:
        return {"success": False, "message": "openpyxl not installed. Run: pip install openpyxl"}

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"success": False, "message": "File is empty or has no data rows"}

    header = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]
    col_map = {name: i for i, name in enumerate(header)}

    required = ["business_name"]
    for req in required:
        if req not in col_map:
            return {"success": False, "message": f"Missing required column: {req}. Found: {', '.join(header)}"}

    created, errors = _process_excel_rows(rows, col_map)

    sb_create_audit_log("BULK_EXCEL_UPLOAD", "admin", "", f"Uploaded {len(created)} offers from {file.filename}")

    return {
        "success": True,
        "message": f"Created {len(created)} offers from {file.filename}",
        "data": {
            "created_count": len(created),
            "error_count": len(errors),
            "errors": errors[:20],
            "total_rows": len(rows) - 1,
        },
    }


@admin_offers_router.get("/admin/offers/upload-template")
def download_upload_template():
    """Download a sample .xlsx template for bulk offer upload."""
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "message": "openpyxl not installed"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Offers"
    headers = [
        "business_name", "title", "description", "business_type",
        "discount_percent", "is_free_item", "address", "offer_url",
        "lat", "lng", "expires_days",
        "source", "original_price", "affiliate_tracking_url", "external_id",
    ]
    ws.append(headers)
    ws.append([
        "Coffee House", "15% off drinks", "Get 15% off any beverage", "cafe",
        15, "no", "123 Main St, Columbus OH", "https://coffeehouse.com/snap",
        39.9612, -82.9988, 30,
        "direct", "", "", "",
    ])
    ws.append([
        "Auto Spa", "Free car wash", "Complimentary full wash", "service",
        100, "yes", "456 Oak Ave, Columbus OH", "",
        39.97, -83.00, 14,
        "direct", "", "", "",
    ])
    ws.append([
        "Groupon Pizza", "50% off large pizza", "Half price large pizza deal", "restaurant",
        50, "no", "789 Elm St, Columbus OH", "https://groupon.com/deals/pizza",
        39.98, -82.99, 60,
        "groupon", 24.99, "https://cj.com/track/pizza123", "GRP-12345",
    ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=snaproad_offers_template.xlsx"},
    )


@admin_offers_router.put("/admin/offers/{offer_id}")
def update_offer(offer_id: str, offer_data: dict):
    success = sb_update_offer(offer_id, offer_data)
    if success:
        return {"success": True, "message": "Offer updated successfully"}
    return {"success": False, "message": "Failed to update offer"}


@admin_offers_router.delete("/admin/offers/{offer_id}")
def delete_offer(offer_id: str):
    success = sb_delete_offer(offer_id)
    if success:
        sb_create_audit_log("OFFER_DELETED", "admin", offer_id, "Deleted offer")
        return {"success": True, "message": "Offer deleted successfully"}
    return {"success": False, "message": "Failed to delete offer"}


@admin_offers_router.get("/admin/export/offers")
def export_offers(format: str = "json"):
    offers = sb_get_offers(status="all")
    if format == "csv" and offers:
        headers = list(offers[0].keys())
        csv_lines = [",".join(headers)] + [",".join(str(o.get(h, "")) for h in headers) for o in offers]
        return {"success": True, "data": "\n".join(csv_lines), "format": "csv", "count": len(offers)}
    return {"success": True, "data": offers, "format": "json", "count": len(offers)}


@admin_users_router.get("/admin/export/users")
def export_users(format: str = "json"):
    users = sb_list_profiles(limit=5000)
    if format == "csv" and users:
        headers = list(users[0].keys())
        csv_lines = [",".join(headers)] + [",".join(str(u.get(h, "")) for h in headers) for u in users]
        return {"success": True, "data": "\n".join(csv_lines), "format": "csv", "count": len(users)}
    return {"success": True, "data": users, "format": "json", "count": len(users)}


@admin_offers_router.post("/admin/import/offers")
def import_offers(import_data: OfferImport):
    imported = 0
    errors = []
    for offer_data in import_data.offers:
        try:
            data = {
                "business_name": offer_data.get("business_name", "Imported"),
                "business_type": offer_data.get("business_type", "retail"),
                "description": offer_data.get("description", ""),
                "title": offer_data.get("title", offer_data.get("description", "Import")[:60]),
                "base_gems": offer_data.get("base_gems", 25),
                "lat": offer_data.get("lat", 39.9612),
                "lng": offer_data.get("lng", -82.9988),
                "is_admin_offer": True,
                "created_by": "admin_import",
                "status": "active",
            }
            result = sb_create_offer(data)
            if result:
                imported += 1
        except Exception as e:
            errors.append(str(e))
    return {"success": True, "message": f"Imported {imported} offers", "data": {"imported": imported, "errors": errors}}


# ==================== GROUPON / CJ AFFILIATE IMPORT ====================

@admin_offers_router.post("/admin/offers/import-groupon")
async def import_groupon_deals(
    area: Annotated[str, Query()] = "Columbus, OH",
    category: Annotated[Optional[str], Query()] = None,
    limit: Annotated[int, Query(ge=1, le=100)] = 20,
):
    """Fetch deals from Groupon via CJ Affiliate API and return a preview list."""
    from services.groupon_service import fetch_groupon_deals, import_deals_to_offers

    raw_deals = await fetch_groupon_deals(area=area, category=category, limit=limit)
    if not raw_deals:
        return {
            "success": True,
            "message": "No deals found (CJ_API_KEY may not be configured yet)",
            "data": {"deals": [], "count": 0},
        }

    offer_previews = import_deals_to_offers(raw_deals)
    return {
        "success": True,
        "message": f"Fetched {len(offer_previews)} deals from Groupon",
        "data": {"deals": offer_previews, "count": len(offer_previews)},
    }


@admin_offers_router.post("/admin/offers/approve-imports")
def approve_imported_deals(deals: list[dict]):
    """
    Receive a list of pre-normalised deal dicts (from import-groupon preview)
    and insert them as active offers.
    """
    created = []
    errors = []
    for idx, deal in enumerate(deals):
        try:
            deal["status"] = "active"
            result = sb_create_offer(deal)
            if result:
                created.append(result)
            else:
                errors.append(f"Deal {idx}: DB insert failed")
        except Exception as e:
            errors.append(f"Deal {idx}: {e}")

    if created:
        sb_create_audit_log(
            "GROUPON_IMPORT_APPROVED", "admin", "",
            f"Approved {len(created)} Groupon deals",
        )

    return {
        "success": True,
        "message": f"Approved {len(created)} of {len(deals)} deals",
        "data": {"created_count": len(created), "error_count": len(errors), "errors": errors[:20]},
    }


# ==================== YELP ENRICHMENT ====================

@admin_offers_router.post("/admin/offers/{offer_id}/enrich-yelp")
async def enrich_offer_with_yelp(offer_id: str):
    """Fetch Yelp rating, reviews and photo for an offer and update it."""
    from services.yelp_service import enrich_offer

    offers = sb_get_offers(status="all")
    offer = next((o for o in offers if str(o.get("id")) == offer_id), None)
    if not offer:
        return {"success": False, "message": "Offer not found"}

    enrichment = await enrich_offer(
        business_name=offer.get("business_name", ""),
        lat=offer.get("lat"),
        lng=offer.get("lng"),
    )
    if not enrichment:
        return {"success": False, "message": "No Yelp data found for this business"}

    success = sb_update_offer(offer_id, enrichment)
    if success:
        sb_create_audit_log("OFFER_YELP_ENRICHED", "admin", offer_id, f"Enriched with Yelp: {enrichment.get('yelp_rating')}★")
        return {"success": True, "message": "Offer enriched with Yelp data", "data": enrichment}
    return {"success": False, "message": "Failed to update offer with Yelp data"}


# ==================== PARTNERS CRUD ====================

@admin_partners_router.get("/admin/partners")
def get_partners(limit: Annotated[int, Query(ge=1, le=100)] = 100):
    data = sb_get_partners(limit=limit)
    return {"success": True, "data": data}


@admin_partners_router.post("/admin/partners")
def create_partner(partner_data: dict):
    partner_data.setdefault("status", "pending")
    partner_data.setdefault("is_approved", False)
    result = sb_create_partner(partner_data)
    if result:
        sb_create_audit_log("PARTNER_CREATED", "admin", result.get("id", ""), f"Created partner: {partner_data.get('business_name', '')}")
        return {"success": True, "message": "Partner created successfully", "data": result}
    return {"success": False, "message": "Failed to create partner"}


@admin_partners_router.put("/admin/partners/{partner_id}")
def update_partner(partner_id: str, partner_data: dict):
    body = dict(partner_data or {})
    if "plan" in body and body.get("plan") is not None:
        body["plan_entitlement_source"] = "admin"
        pl = str(body.get("plan") or "").strip().lower()
        # Admin-assigned partner tiers should unlock the portal without Stripe checkout.
        if pl in ("starter", "growth", "enterprise", "internal") and "subscription_status" not in body:
            body["subscription_status"] = "active"
    success = sb_update_partner(partner_id, body)
    if success:
        return {"success": True, "message": "Partner updated successfully"}
    return {"success": False, "message": "Failed to update partner"}


@admin_partners_router.delete("/admin/partners/{partner_id}")
def delete_partner(partner_id: str):
    success = sb_delete_partner(partner_id)
    if success:
        sb_create_audit_log("PARTNER_DELETED", "admin", partner_id, "Deleted partner")
        return {"success": True, "message": "Partner deleted successfully"}
    return {"success": False, "message": "Failed to delete partner"}


@admin_partners_router.post("/admin/partners/{partner_id}/approve")
def approve_partner(partner_id: str):
    success = sb_update_partner(partner_id, {"status": "active", "is_approved": True})
    if success:
        sb_create_audit_log("PARTNER_APPROVED", "admin", partner_id, "Approved partner")
        return {"success": True, "message": "Partner approved successfully"}
    return {"success": False, "message": "Failed to approve partner"}


@admin_partners_router.post("/admin/partners/{partner_id}/suspend")
def suspend_partner(partner_id: str):
    success = sb_update_partner(partner_id, {"status": "suspended"})
    if success:
        sb_create_audit_log("PARTNER_SUSPENDED", "admin", partner_id, "Suspended partner")
        return {"success": True, "message": "Partner suspended successfully"}
    return {"success": False, "message": "Failed to suspend partner"}


# ==================== CAMPAIGNS CRUD ====================

@admin_campaigns_router.get("/admin/campaigns")
def get_campaigns(limit: Annotated[int, Query(ge=1, le=100)] = 100):
    data = sb_get_campaigns(limit=limit)
    return {"success": True, "data": data}


@admin_campaigns_router.post("/admin/campaigns")
def create_campaign(campaign_data: dict):
    campaign_data.setdefault("status", "draft")
    result = sb_create_campaign(campaign_data)
    if result:
        return {"success": True, "message": "Campaign created successfully", "data": result}
    return {"success": False, "message": "Failed to create campaign"}


@admin_campaigns_router.put("/admin/campaigns/{campaign_id}")
def update_campaign(campaign_id: str, campaign_data: dict):
    success = sb_update_campaign(campaign_id, campaign_data)
    if success:
        return {"success": True, "message": "Campaign updated successfully"}
    return {"success": False, "message": "Failed to update campaign"}


@admin_campaigns_router.delete("/admin/campaigns/{campaign_id}")
def delete_campaign(campaign_id: str):
    success = sb_delete_campaign(campaign_id)
    if success:
        return {"success": True, "message": "Campaign deleted successfully"}
    return {"success": False, "message": "Failed to delete campaign"}


@admin_campaigns_router.post("/admin/campaigns/{campaign_id}/activate")
def activate_campaign(campaign_id: str):
    success = sb_update_campaign(campaign_id, {"status": "active"})
    if success:
        return {"success": True, "message": "Campaign activated successfully"}
    return {"success": False, "message": "Failed to activate campaign"}


# ==================== REWARDS CRUD ====================

@admin_rewards_router.get("/admin/rewards")
def get_rewards(limit: Annotated[int, Query(ge=1, le=100)] = 100):
    data = sb_get_rewards(limit=limit)
    return {"success": True, "data": data}


@admin_rewards_router.post("/admin/rewards")
def create_reward(reward_data: dict):
    reward_data.setdefault("status", "active")
    result = sb_create_reward(reward_data)
    if result:
        return {"success": True, "message": "Reward created successfully", "data": result}
    return {"success": False, "message": "Failed to create reward"}


@admin_rewards_router.put("/admin/rewards/{reward_id}")
def update_reward(reward_id: str, reward_data: dict):
    success = sb_update_reward(reward_id, reward_data)
    if success:
        return {"success": True, "message": "Reward updated successfully"}
    return {"success": False, "message": "Failed to update reward"}


@admin_rewards_router.delete("/admin/rewards/{reward_id}")
def delete_reward(reward_id: str):
    success = sb_delete_reward(reward_id)
    if success:
        return {"success": True, "message": "Reward deleted successfully"}
    return {"success": False, "message": "Failed to delete reward"}


@admin_rewards_router.post("/admin/rewards/{reward_id}/claim")
def claim_reward(reward_id: str, user_data: dict):
    rewards = sb_get_rewards()
    reward = next((r for r in rewards if r["id"] == reward_id), None)
    if not reward:
        return {"success": False, "message": "Reward not found"}
    if reward.get("claimed", 0) >= reward.get("total", 0):
        return {"success": False, "message": "Reward out of stock"}
    sb_update_reward(reward_id, {"claimed": reward.get("claimed", 0) + 1})
    return {"success": True, "message": "Reward claimed successfully"}


# ==================== USERS CRUD ====================

def _effective_partner_row_id_for_profile(u: dict) -> Optional[str]:
    """Partner row id for admin: explicit profiles.partner_id, or self id when role is partner
    (partner signup sets public.partners.id = auth user id; partner_id FK is often unset)."""
    pid = str(u.get("partner_id") or "").strip()
    if pid:
        return pid
    if str(u.get("role") or "").strip().lower() == "partner":
        uid = str(u.get("id") or "").strip()
        if uid:
            return uid
    return None


@admin_users_router.get("/admin/users")
def get_users(limit: Annotated[int, Query(ge=1, le=2000)] = 500):
    data = sb_list_profiles(limit=limit)
    pids = []
    for u in data:
        ep = _effective_partner_row_id_for_profile(u)
        if ep:
            pids.append(ep)
    pmap = sb_get_partners_plan_fields_by_ids(list(dict.fromkeys(pids))) if pids else {}
    for u in data:
        ep = _effective_partner_row_id_for_profile(u)
        if not ep:
            continue
        pr = pmap.get(ep)
        if not pr:
            continue
        u["partner_plan"] = pr.get("plan")
        u["partner_subscription_status"] = pr.get("subscription_status")
        u["partner_plan_entitlement_source"] = pr.get("plan_entitlement_source")
        u["partner_promotion_access_until"] = pr.get("promotion_access_until")
        u["partner_is_internal_complimentary"] = bool(pr.get("is_internal_complimentary"))
    return {"success": True, "source": "supabase", "data": data, "total": len(data)}


@admin_users_router.put("/admin/users/{user_id}")
def update_user(user_id: str, user_data: dict):
    body = dict(user_data or {})
    if "plan" in body or "is_premium" in body:
        pl = body.get("plan")
        pls = str(pl).strip().lower() if pl is not None and str(pl).strip() != "" else None
        ip = body.get("is_premium")
        upgrading = ip is True or pls in ("premium", "family")
        downgrading = ip is False or pls == "basic"
        if upgrading:
            body["plan_entitlement_source"] = "admin"
            if pls in ("premium", "family"):
                body["plan"] = pls
            if ip is True and not pls:
                body.setdefault("plan", "premium")
            body["is_premium"] = True
            if body.get("gem_multiplier") is None:
                body["gem_multiplier"] = 2
        elif downgrading:
            body["plan_entitlement_source"] = None
            body["plan"] = "basic"
            body["is_premium"] = False
            if body.get("gem_multiplier") is None:
                body["gem_multiplier"] = 1
    success = sb_update_profile(user_id, body)
    if success:
        # Users tab edits driver profile plan only. Partner portal entitlement lives on `partners`.
        # If this account is linked to a partner row, keep subscription active when admin changes plan tier.
        try:
            prof = sb_get_profile_raw(user_id) or {}
            pid = str(prof.get("partner_id") or "").strip()
            if not pid and str(prof.get("role") or "").strip().lower() == "partner":
                pid = str(user_id).strip()
            if pid and ("plan" in body or "is_premium" in body):
                prow = sb_get_partner(pid)
                if prow:
                    pu: dict = {
                        "subscription_status": "active",
                        "plan_entitlement_source": "admin",
                    }
                    pl = str(prow.get("plan") or "").strip().lower()
                    if pl in ("", "unselected", "none"):
                        pu["plan"] = "starter"
                    sb_update_partner(pid, pu)
        except Exception as e:
            logger.warning("partner sync after admin user update failed: %s", e)
        return {"success": True, "message": "User updated successfully"}
    return {"success": False, "message": "Failed to update user"}


@admin_users_router.delete("/admin/users/{user_id}")
def delete_user(user_id: str):
    success = sb_delete_profile(user_id)
    if success:
        sb_create_audit_log("USER_DELETED", "admin", user_id, "Deleted user")
        return {"success": True, "message": "User deleted successfully"}
    return {"success": False, "message": "Failed to delete user"}


@admin_users_router.post("/admin/users/{user_id}/suspend")
def suspend_user(user_id: str):
    success = sb_suspend_profile(user_id)
    if success:
        sb_create_audit_log("USER_SUSPENDED", "admin", user_id, "Suspended user")
        return {"success": True, "message": "User suspended successfully"}
    return {"success": False, "message": "Failed to suspend user"}


@admin_users_router.post("/admin/users/{user_id}/activate")
def activate_user(user_id: str):
    success = sb_activate_profile(user_id)
    if success:
        return {"success": True, "message": "User activated successfully"}
    return {"success": False, "message": "Failed to activate user"}


_USER_PROMO_PLANS = frozenset({"premium", "family"})
_PARTNER_PROMO_PLANS = frozenset({"starter", "growth", "enterprise"})


class GrantPromotionBody(BaseModel):
    user_ids: List[str] = Field(default_factory=list)
    partner_ids: List[str] = Field(default_factory=list)
    days: int = Field(default=30, ge=1, le=730)
    plan: str = Field(default="premium", min_length=1, max_length=32)
    send_email: bool = False


@admin_users_router.post("/admin/promotions/grant")
def grant_promotions(body: GrantPromotionBody, admin: AdminUser):
    """
    Extend or start time-boxed access for drivers (premium/family benefits) or partners
    (subscription treated as active). Optionally notify by email (Resend).
    """
    uids = [str(x).strip() for x in (body.user_ids or []) if str(x).strip()]
    pids = [str(x).strip() for x in (body.partner_ids or []) if str(x).strip()]
    if not uids and not pids:
        return {"success": False, "message": "Select at least one user or partner"}
    uids = uids[:300]
    pids = pids[:300]

    ref = f"PROMO-{secrets.token_hex(4).upper()}"
    actor = str(admin.get("id") or admin.get("user_id") or "admin")
    user_plan = str(body.plan or "premium").strip().lower()
    if user_plan not in _USER_PROMO_PLANS:
        user_plan = "premium"
    partner_plan = str(body.plan or "growth").strip().lower()
    if partner_plan not in _PARTNER_PROMO_PLANS:
        partner_plan = "growth"

    updated_users = 0
    updated_partners = 0
    emails_sent = 0
    email_errors: list[str] = []

    for uid in uids:
        raw_until = sb_get_profile_promotion_until_raw(uid)
        until_iso = compute_extended_promotion_until_iso(raw_until, body.days)
        if sb_update_profile(uid, {"promotion_access_until": until_iso, "promotion_plan": user_plan}):
            updated_users += 1
        if body.send_email:
            row = sb_get_profile(uid)
            to = (row or {}).get("email")
            name = (row or {}).get("name") or "there"
            if to:
                until_disp = until_iso[:10] if len(until_iso) >= 10 else until_iso
                html = promotion_email_html(
                    str(name),
                    is_partner=False,
                    plan_label=user_plan.title(),
                    until_display=until_disp,
                    reference=ref,
                )
                ok, err = send_html_email(
                    str(to),
                    f"SnapRoad: {body.days} days of {user_plan.title()} access",
                    html,
                )
                if ok:
                    emails_sent += 1
                elif err:
                    email_errors.append(f"{to}: {err}")

    for pid in pids:
        raw_until = sb_get_partner_promotion_until_raw(pid)
        until_iso = compute_extended_promotion_until_iso(raw_until, body.days)
        if sb_update_partner(pid, {"promotion_access_until": until_iso, "promotion_plan": partner_plan}):
            updated_partners += 1
        if body.send_email:
            prow = sb_get_partner(pid)
            to = (prow or {}).get("email")
            name = (prow or {}).get("business_name") or "there"
            if to:
                until_disp = until_iso[:10] if len(until_iso) >= 10 else until_iso
                html = promotion_email_html(
                    str(name),
                    is_partner=True,
                    plan_label=partner_plan.title(),
                    until_display=until_disp,
                    reference=ref,
                )
                ok, err = send_html_email(
                    str(to),
                    f"SnapRoad Partner: {body.days} days complimentary {partner_plan.title()}",
                    html,
                )
                if ok:
                    emails_sent += 1
                elif err:
                    email_errors.append(f"{to}: {err}")

    try:
        sb_create_audit_log(
            "PROMOTION_GRANTED",
            actor,
            ref,
            f"users={updated_users} partners={updated_partners} days={body.days} email={body.send_email}",
        )
    except Exception:
        logger.debug("audit log for PROMOTION_GRANTED failed (non-critical)")

    return {
        "success": True,
        "data": {
            "reference": ref,
            "updated_users": updated_users,
            "updated_partners": updated_partners,
            "emails_sent": emails_sent,
            "email_errors": email_errors[:12],
        },
    }


# ==================== PRICING ====================

@admin_pricing_router.get("/admin/pricing")
def get_admin_pricing():
    settings = sb_get_settings()
    pricing = settings.get("pricing", {
        "founders_price": 7.99,
        "public_price": 7.99,
        "is_founders_active": True,
    })
    return {"success": True, "data": pricing}


@admin_pricing_router.post("/admin/pricing")
def update_admin_pricing(update: PricingUpdate):
    pricing = {}
    if update.founders_price is not None:
        pricing["founders_price"] = update.founders_price
    if update.public_price is not None:
        pricing["public_price"] = update.public_price
    if update.is_founders_active is not None:
        pricing["is_founders_active"] = update.is_founders_active
    sb_update_setting("pricing", pricing)
    return {"success": True, "message": "Pricing updated", "data": pricing}


# ==================== BOOSTS ====================

@admin_boosts_router.post("/boosts/calculate")
def calculate_boost_cost(calc: BoostCalculate):
    duration_cost = BOOST_PRICING_ADMIN["base_daily_cost"] + max(0, calc.duration_days - 1) * BOOST_PRICING_ADMIN["additional_day_cost"]
    reach_increments = calc.reach_target // BOOST_PRICING_ADMIN["reach_increment"]
    reach_cost = BOOST_PRICING_ADMIN["base_reach_cost"] + max(0, reach_increments - 1) * BOOST_PRICING_ADMIN["reach_increment_cost"]
    return {"success": True, "data": {
        "duration_days": calc.duration_days,
        "reach_target": calc.reach_target,
        "duration_cost": duration_cost,
        "reach_cost": reach_cost,
        "total_cost": duration_cost + reach_cost,
    }}


@admin_boosts_router.post("/boosts/create")
def create_boost(boost: BoostCreate):
    calc = calculate_boost_cost(BoostCalculate(duration_days=boost.duration_days, reach_target=boost.reach_target))
    data = {
        "offer_id": boost.offer_id,
        "partner_id": boost.business_id or None,
        "budget": calc["data"]["total_cost"],
        "duration_days": boost.duration_days,
        "target_radius_miles": boost.reach_target,
        "status": "active",
        "ends_at": (datetime.now() + timedelta(days=boost.duration_days)).isoformat(),
    }
    result = sb_create_boost(data)
    if result:
        return {"success": True, "message": f"Boost created! ${calc['data']['total_cost']} charged.", "data": result}
    return {"success": False, "message": "Failed to create boost"}


@admin_boosts_router.get("/boosts")
def get_boosts(
    partner_id: Annotated[Optional[str], Query()] = None,
    limit: Annotated[int, Query(ge=1, le=100)] = 100,
):
    data = sb_get_boosts(partner_id=partner_id)[:limit]
    return {"success": True, "data": data}


@admin_boosts_router.get("/boosts/{boost_id}")
def get_boost(boost_id: str):
    boosts = sb_get_boosts()
    boost = next((b for b in boosts if b["id"] == boost_id), None)
    if not boost:
        return {"success": False, "message": "Boost not found"}
    return {"success": True, "data": boost}


@admin_boosts_router.delete("/boosts/{boost_id}")
def cancel_boost(boost_id: str):
    success = sb_cancel_boost(boost_id)
    if success:
        return {"success": True, "message": "Boost cancelled"}
    return {"success": False, "message": "Failed to cancel boost"}


# ==================== ANALYTICS TRACKING ====================

@admin_ops_router.post("/analytics/track")
def track_analytics_event(event: AnalyticsEvent):
    offers = sb_get_offers(limit=500)
    offer = next((o for o in offers if str(o.get("id")) == str(event.offer_id)), None)
    if not offer:
        return {"success": False, "message": "Offer not found"}
    tracked = record_offer_event(
        offer=offer,
        event_type=event.event_type,
        partner_id=offer.get("partner_id"),
        user_id=None,
        lat=(event.user_location or {}).get("lat") if isinstance(event.user_location, dict) else None,
        lng=(event.user_location or {}).get("lng") if isinstance(event.user_location, dict) else None,
    )
    return {"success": True, "message": "Event tracked", "data": tracked}


@admin_ops_router.get("/analytics/dashboard")
def get_analytics_dashboard(
    business_id: Annotated[str, Query()] = "default_business",
    days: Annotated[int, Query(ge=1, le=366)] = 7,
):
    redemption_stats = sb_get_redemption_stats()
    offer_summary = summarize_offer_analytics(limit=200)
    return {
        "success": True,
        "data": {
            "summary": {
                "total_views": sum(int(row.get("views") or 0) for row in offer_summary),
                "total_clicks": sum(int(row.get("visits") or 0) for row in offer_summary),
                "total_redemptions": redemption_stats.get("total", 0),
                "total_revenue": sum(int(row.get("redemptions") or 0) for row in offer_summary),
            },
            "offers": offer_summary[:50],
            "window": {"business_id": business_id, "days": days},
        },
    }


@admin_realtime_router.get("/admin/fees/summary")
def get_admin_fees_summary(month_year: Optional[str] = None):
    return {"success": True, "data": list_monthly_fee_summaries(month_year)}


@admin_realtime_router.get("/admin/fees/partner/{partner_id}")
def get_admin_partner_fee_history(partner_id: str, limit: Annotated[int, Query(ge=1, le=36)] = 12):
    return {"success": True, "data": get_partner_fee_history(partner_id, limit)}


@admin_offers_router.get("/admin/offers/analytics")
def get_admin_offer_analytics(limit: Annotated[int, Query(ge=1, le=500)] = 200):
    return {"success": True, "data": summarize_offer_analytics(limit)}


@admin_realtime_router.get("/admin/realtime/summary")
def get_admin_realtime_summary():
    return {"success": True, "data": today_realtime_summary()}


@admin_realtime_router.get("/admin/realtime/feed")
def get_admin_realtime_feed(limit: Annotated[int, Query(ge=1, le=200)] = 50):
    return {"success": True, "data": recent_offer_feed(limit)}


@admin_realtime_router.get("/admin/realtime/map-data")
def get_admin_realtime_map_data():
    return {"success": True, "data": map_data_for_day()}


# ==================== SUPABASE STATUS ====================

@admin_platform_router.get("/admin/supabase/status")
def get_supabase_status():
    status = test_connection()
    return {"success": True, "data": status}


@admin_platform_router.get("/admin/events")
def get_admin_events(limit: Annotated[int, Query(ge=1, le=100)] = 100):
    challenges = sb_get_challenges()[:limit]
    return {"success": True, "data": challenges}


# ==================== PHOTO REPORT MODERATION ====================


def _photo_original_signed_url(supabase, storage_path: str) -> Optional[str]:
    from services.photo_report_processing import PRIVATE_ORIGINALS_BUCKET

    try:
        r = supabase.storage.from_(PRIVATE_ORIGINALS_BUCKET).create_signed_url(storage_path, 3600)
        if isinstance(r, dict):
            return r.get("signedURL") or r.get("signed_url") or r.get("signedUrl")
        su = getattr(r, "signed_url", None) or getattr(r, "signedURL", None)
        return str(su) if su else None
    except Exception as e:
        logger.warning("signed URL for original failed: %s", e)
        return None


@admin_photo_reports_router.get("/admin/photo-reports/pending")
def admin_photo_reports_pending(limit: Annotated[int, Query(ge=1, le=200)] = 50):
    from database import get_supabase

    supabase = get_supabase()
    res = (
        supabase.table("incident_photos")
        .select("*")
        .eq("moderation_status", "pending_review")
        .order("created_at", desc=True)
        .limit(limit)
        .execute()
    )
    reports = []
    for row in res.data or []:
        item = dict(row)
        opath = item.get("original_storage_path")
        if opath:
            item["original_signed_url"] = _photo_original_signed_url(supabase, opath)
        reports.append(item)
    return {"success": True, "data": {"reports": reports, "total": len(reports)}}


@admin_photo_reports_router.post("/admin/photo-reports/{report_id}/approve")
async def admin_photo_report_approve(report_id: str):
    import uuid

    from database import get_supabase

    from services.photo_report_processing import (
        PUBLIC_PHOTO_BUCKET,
        PRIVATE_ORIGINALS_BUCKET,
        analyze_and_prepare_public_bytes,
        heavy_blur_full_jpeg,
    )

    supabase = get_supabase()
    fetch = (
        supabase.table("incident_photos")
        .select("*")
        .eq("id", report_id)
        .eq("moderation_status", "pending_review")
        .limit(1)
        .execute()
    )
    if not fetch.data:
        return {"success": False, "message": "Report not found or not pending"}
    row = fetch.data[0]
    opath = row.get("original_storage_path")
    uid = row.get("user_id") or "unknown"
    if not opath:
        return {"success": False, "message": "No private original path"}
    try:
        raw = supabase.storage.from_(PRIVATE_ORIGINALS_BUCKET).download(opath)
        if not raw:
            return {"success": False, "message": "Could not download original"}
    except Exception as e:
        logger.exception("download original: %s", e)
        return {"success": False, "message": "Could not download original"}

    prep = await analyze_and_prepare_public_bytes(raw)
    needs_pii = prep["needs_pii"]
    regions_blurred = prep["regions_blurred"]
    analysis_ok = prep["analysis_ok"]

    if needs_pii and regions_blurred > 0:
        public_bytes = prep["blurred_jpeg"]
        blur_applied = True
    elif (not analysis_ok) or (needs_pii and regions_blurred == 0):
        public_bytes = heavy_blur_full_jpeg(raw)
        blur_applied = True
    else:
        public_bytes = prep["base_jpeg"]
        blur_applied = False

    file_path = f"{uid}/{uuid.uuid4().hex}_approved.jpg"
    try:
        supabase.storage.from_(PUBLIC_PHOTO_BUCKET).upload(
            file_path,
            public_bytes,
            file_options={"content-type": "image/jpeg"},
        )
        photo_url = supabase.storage.from_(PUBLIC_PHOTO_BUCKET).get_public_url(file_path)
    except Exception as e:
        logger.exception("public upload on approve: %s", e)
        return {"success": False, "message": "Public storage upload failed"}

    try:
        supabase.table("incident_photos").update(
            {
                "photo_url": photo_url,
                "thumbnail_url": photo_url,
                "moderation_status": "active",
                "blur_applied": blur_applied,
                "needs_admin_review": False,
                "original_storage_path": None,
                "review_notes": None,
            }
        ).eq("id", report_id).execute()
    except Exception as e:
        logger.exception("db update approve: %s", e)
        return {"success": False, "message": "Database update failed"}

    try:
        supabase.storage.from_(PRIVATE_ORIGINALS_BUCKET).remove([opath])
    except Exception:
        logger.warning("could not delete private original %s", opath)

    return {"success": True, "message": "Published blurred image", "data": {"photo_url": photo_url}}


@admin_photo_reports_router.post("/admin/photo-reports/{report_id}/reject")
def admin_photo_report_reject(
    report_id: str,
    body: Annotated[AdminPhotoRejectBody, Body(default_factory=AdminPhotoRejectBody)],
):
    from database import get_supabase

    from services.photo_report_processing import PRIVATE_ORIGINALS_BUCKET

    notes = body.review_notes
    supabase = get_supabase()
    fetch = (
        supabase.table("incident_photos")
        .select("id,original_storage_path")
        .eq("id", report_id)
        .eq("moderation_status", "pending_review")
        .limit(1)
        .execute()
    )
    if not fetch.data:
        return {"success": False, "message": "Report not found or not pending"}
    opath = fetch.data[0].get("original_storage_path")
    if opath:
        try:
            supabase.storage.from_(PRIVATE_ORIGINALS_BUCKET).remove([opath])
        except Exception:
            logger.warning("could not remove private object %s", opath)
    try:
        supabase.table("incident_photos").update(
            {
                "moderation_status": "rejected",
                "photo_url": None,
                "thumbnail_url": None,
                "needs_admin_review": False,
                "original_storage_path": None,
                "review_notes": str(notes)[:2000] if notes else None,
            }
        ).eq("id", report_id).execute()
    except Exception as e:
        logger.exception("reject update: %s", e)
        return {"success": False, "message": "Database update failed"}
    return {"success": True, "message": "Report rejected"}


from routes.admin_metrics import admin_metrics_router  # noqa: E402
from routes.admin_link_offers import admin_link_offers_router  # noqa: E402

router.include_router(admin_platform_router)
router.include_router(admin_concerns_router)
router.include_router(admin_config_router)
router.include_router(admin_incidents_router)
router.include_router(admin_offers_router)
router.include_router(admin_link_offers_router)
router.include_router(admin_partners_router)
router.include_router(admin_campaigns_router)
router.include_router(admin_rewards_router)
router.include_router(admin_users_router)
router.include_router(admin_pricing_router)
router.include_router(admin_boosts_router)
router.include_router(admin_realtime_router)
router.include_router(admin_photo_reports_router)
router.include_router(admin_ops_router)
router.include_router(admin_metrics_router)
